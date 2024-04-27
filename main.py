import csv
import sqlite3
from ui import Ui
from veri_tabani import Kitap
from openpyxl import Workbook


def kitap_ekle(kitap):
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    cursor.execute("INSERT INTO kitaplar (ISBN, KITAP_ADI, YAZAR, YAYIN_TARIHI, TUR, OZET) VALUES (?, ?, ?, ?, ?, ?)",
                   (kitap.ISBN, kitap.KITAP_ADI, kitap.YAZAR, kitap.YAYIN_TARIHI, kitap.TUR, kitap.OZET))

    print("kitap başarı ile eklendi")
    baglaninal_database.commit()
    baglaninal_database.close()


def kitap_ara():
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()

    menu1 = Ui(["isbn arama", "kitap adı arama", "yazar adi arama"])
    Ui.Show_Menu(menu1)

    secim = int(input("aramak istediğiniz şeyi seçiniz: "))

    if secim == 1:
        deger1 = input("aranacak parametreleri giriniz: ")
        cursor.execute("SELECT * FROM Kitaplar WHERE ISBN LIKE ?", ('%' + deger1 + '%',))
        print(cursor.fetchall())
        baglaninal_database.close()

    elif secim == 2:
        deger2 = input("aranacak parametreleri giriniz: ")
        cursor.execute("SELECT * FROM Kitaplar WHERE ISBN LIKE ?", ('%' + deger2 + '%',))
        print(cursor.fetchall())
        baglaninal_database.close()

    elif secim == 3:
        deger3 = input("aranacak parametreleri giriniz: ")
        cursor.execute("SELECT * FROM Kitaplar WHERE ISBN LIKE ?", ('%' + deger3 + '%',))
        print(cursor.fetchall())
        baglaninal_database.close()

    else:
        print("hata")


def kitap_sil():
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()

    deger = input("Silmek istediğiniz kitabın ISBN'ini giriniz: ")
    cursor.execute("SELECT * FROM kitaplar WHERE ISBN=?", (deger,))
    kitap = cursor.fetchone()

    if kitap:
        deger1 = input(f"{kitap} bu kitabı gerçekten silmek istiyor musunuz? (1 evet, 0 hayır): ")
        if deger1 == '1':
            cursor.execute("DELETE FROM kitaplar WHERE ISBN=?", (deger,))
            baglaninal_database.commit()
            baglaninal_database.close()
            print("Kitap başarıyla silindi.")
        elif deger1 == '0':
            print("İşlem iptal edildi.")
        else:
            print("Hatalı giriş.")
    else:
        print("Girilen ISBN'e sahip kitap bulunamadı.")


def kitap_listele():
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    cursor.execute("SELECT * FROM Kitaplar")
    print(cursor.fetchall())
    baglaninal_database.close()


def kitap_odunc_alma():
    kitap_isbn = input("ödünç alınacak kitabın isbn sini giriniz: ")
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    try:
        cursor.execute(f"SELECT ODUNC_ALINDI_MI FROM Kitaplar WHERE ISBN={kitap_isbn}")
        if cursor.fetchone()[0]:
            print("kitap kütüphanede değil")
        else:
            cursor.execute(f"UPDATE Kitaplar SET ODUNC_ALINDI_MI = 1 WHERE ISBN ={kitap_isbn}")
            baglaninal_database.commit()
        baglaninal_database.close()

    except Exception as e:
        print(e)


def kitap_odunc_verme():
    kitap_isbn = input("ödünç alınacak kitabın isbn sini giriniz: ")
    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    try:
        cursor.execute(f"UPDATE Kitaplar SET ODUNC_ALINDI_MI = 1 WHERE ISBN ={kitap_isbn}")
        baglaninal_database.commit()
        baglaninal_database.close()

    except Exception as e:
        print(e)


def excel_yedegi_alma():
    wb = Workbook()
    ws = wb.active

    ws["A1"] = "ISBN"
    ws["B1"] = "KITAP_ADI"
    ws["C1"] = "YAZAR"
    ws["D1"] = "YAYIN_TARIHI"
    ws["E1"] = "TUR"
    ws["F1"] = "OZET"
    ws["G1"] = "ODUNC_ALINDI_MI"

    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    cursor.execute("SELECT * FROM kitaplar ")

    for satir in cursor.fetchall():
        ws.append(satir)

    wb.save("dosya.xlsx")


def csv_yedegi_alma():
    with open("dosya.csv", "w", newline="") as dosya:
        baglaninal_database = sqlite3.connect('database.sqlite')
        cursor = baglaninal_database.cursor()
        cursor.execute("SELECT * FROM kitaplar ")

        csv_writer = csv.writer(dosya)
        csv_writer.writerow(["ISBN", "KITAP_ADI", "YAZAR", "YAYIN_TARIHI", "TUR", "OZET", "ODUNC_ALINDI_MI"])
        for satir in cursor.fetchall():
            csv_writer.writerow([satir])


while True:
    menu = Ui(["arama",
               "ekleme",
               "listeleme",
               "silme",
               "ödünç alma",
               "ödünç verme",
               "excele yedek alma",
               "csv'ye kayıt alma",
               "çıkış"])
    Ui.Show_Menu(menu)

    secim1 = int(input("lütfen seçim yapınız: "))

    if secim1 == 1:
        kitap_ara()

    elif secim1 == 2:
        isbn = input("isbn bilgisini giriniz: ")
        kitap_adi = input("kitap adi bilgisini giriniz: ")
        yazar = input("yazar bilgisini giriniz: ")
        yayin_tarihi = input("yayin tarihi bilgisini giriniz: ")
        tur = input("tür bilgisini giriniz: ")
        ozet = input("ozet bilgisini giriniz: ")
        kitap = Kitap(isbn, kitap_adi, yazar, yayin_tarihi, tur, ozet)

        kitap_ekle(kitap)

    elif secim1 == 3:
        kitap_listele()

    elif secim1 == 4:
        kitap_sil()

    elif secim1 == 5:
        kitap_odunc_alma()

    elif secim1 == 6:
        kitap_odunc_verme()

    elif secim1 == 7:
        excel_yedegi_alma()

    elif secim1 == 8:
        csv_yedegi_alma()

    elif secim1 == 9:
        print("çıkış yapılıyor")
        break

    else:
        print("hata")


