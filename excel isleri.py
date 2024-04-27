import sqlite3
from openpyxl import Workbook


def excel_isleri():
    wb = Workbook()

    ws = wb.active

    ws["A1"] = "ISBN"
    ws["B1"] = "KITAP_ADI"
    ws["C1"] = "YAZAR"
    ws["D1"] = "YAYIN_TARIHI"
    ws["E1"] = "TUR"
    ws["F1"] = "OZET"
    ws["G1"] = "ODUNC_ALINDI_MI"

    baglaninal_database = sqlite3.connect('database.sqlite')
    cursor = baglaninal_database.cursor()
    cursor.execute("SELECT * FROM kitaplar ")

    for satir in cursor.fetchall():
        ws.append(satir)

    wb.save("dosya.xlsx")
